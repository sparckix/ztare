"""Read-only extraction of contribution-capacity evidence from the household budget."""

from __future__ import annotations

import hashlib
from pathlib import Path
import re
import statistics
from typing import Any

from openpyxl import load_workbook

from ztare.common.equivariance import stable_sha256

from .contracts import require_finite


HOUSEHOLD_BUDGET_EVIDENCE_SCHEMA = "jaggedthoughts-household-budget-evidence-v1"
_SHEET = "5-Year Plan w Chloe"
_COLUMNS = range(2, 7)


def _numbers(sheet: Any, row: int) -> list[float]:
    return [require_finite(sheet.cell(row, column).value, f"{_SHEET}!{row}:{column}")
            for column in _COLUMNS]


def compile_household_budget_evidence(path: str | Path, *, source_id: str) -> dict[str, Any]:
    """Extract audited annual savings without modifying or trusting broken totals."""
    source = Path(path).expanduser().resolve()
    workbook_sha256 = hashlib.sha256(source.read_bytes()).hexdigest()
    formulas_book = load_workbook(source, read_only=True, data_only=False)
    values_book = load_workbook(source, read_only=True, data_only=True)
    if _SHEET not in formulas_book.sheetnames or _SHEET not in values_book.sheetnames:
        raise ValueError(f"household budget lacks required sheet: {_SHEET}")
    formulas, values = formulas_book[_SHEET], values_book[_SHEET]
    years = []
    for column in _COLUMNS:
        match = re.search(r"\b(20\d{2})\b", str(values.cell(12, column).value or ""))
        if not match:
            raise ValueError("household budget year headers must contain a four-digit year")
        years.append(int(match.group(1)))
    if years != sorted(years) or len(set(years)) != len(years):
        raise ValueError("household budget years must be unique and increasing")

    net_revenue = _numbers(values, 19)
    expense_rows = (29, 37, 55, 58, 65, 69)
    component_expenses = [_numbers(values, row) for row in expense_rows]
    medical = [
        require_finite(values.cell(41, column).value, "medical insurance")
        + require_finite(values.cell(42, column).value, "other medical")
        for column in _COLUMNS
    ]
    corrected_expenses = [
        sum(row[index] for row in component_expenses) + medical[index]
        for index in range(len(years))
    ]
    corrected_savings = [net_revenue[index] - corrected_expenses[index]
                         for index in range(len(years))]
    workbook_savings = _numbers(values, 78)
    financing_double_count = all(
        str(formulas.cell(70, column).value or "").find(f"{formulas.cell(65, column).coordinate}") >= 0
        for column in _COLUMNS
    )
    medical_percentage_in_total = all(
        f"{formulas.cell(38, column).coordinate}:" in str(formulas.cell(43, column).value or "")
        for column in _COLUMNS
    )
    full_year_savings = corrected_savings[1:]
    rows = [{
        "year": year,
        "net_revenue": net_revenue[index],
        "workbook_reported_savings": workbook_savings[index],
        "component_recomputed_savings": corrected_savings[index],
        "formula_defect_impact": corrected_savings[index] - workbook_savings[index],
    } for index, year in enumerate(years)]
    body = {
        "schema": HOUSEHOLD_BUDGET_EVIDENCE_SCHEMA,
        "source_id": str(source_id), "workbook_sha256": workbook_sha256,
        "sheet": _SHEET, "years": years, "annual_rows": rows,
        "contribution_capacity_summary": {
            "full_year_min": min(full_year_savings),
            "full_year_median": statistics.median(full_year_savings),
            "full_year_max": max(full_year_savings),
            "default_scenario_contribution": round(
                statistics.median(full_year_savings) / 1_000
            ) * 1_000,
            "basis": "component_recomputed_savings_2027_2030",
            "operator_confirmed": False,
        },
        "checks": {
            "financing_includes_dependent_care": financing_double_count,
            "medical_total_includes_percentage_row": medical_percentage_in_total,
            "workbook_totals_admitted": False,
            "component_recomputation_admitted_for_scenario_default": True,
        },
        "boundary": (
            "Values are a read-only planning extract. Component recomputation removes the "
            "identified formula overlaps; it does not alter the workbook or confirm future savings."
        ),
        "authority": "private_planning_evidence_only",
        "capital_authority": False,
    }
    return {**body, "budget_evidence_sha256": stable_sha256(body)}


__all__ = ["HOUSEHOLD_BUDGET_EVIDENCE_SCHEMA", "compile_household_budget_evidence"]
